from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = 'waterproof-construction-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///waterproof.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['EXPORT_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')

db = SQLAlchemy(app)

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)

# Database Models
class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    company = db.Column(db.String(100))
    role = db.Column(db.String(50))  # Supplier/Client/Contractor/Other
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.Text)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    location = db.Column(db.String(200))
    contact_person = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    unit = db.Column(db.String(20))  # bag/liter/sqm/pcs/roll/set
    category = db.Column(db.String(50))  # Membrane/Coating/Primer/Sealant/Adhesive/Accessory
    brand = db.Column(db.String(100))
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class PriceEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    price = db.Column(db.Float, nullable=False)
    date_checked = db.Column(db.Date, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    product = db.relationship('Product', backref='price_entries')
    store = db.relationship('Store', backref='price_entries')

# Create tables
with app.app_context():
    db.create_all()

# Routes
@app.route('/')
def index():
    total_contacts = Contact.query.count()
    total_stores = Store.query.count()
    total_products = Product.query.count()
    total_price_entries = PriceEntry.query.count()
    
    recent_prices = PriceEntry.query.order_by(PriceEntry.date_checked.desc()).limit(10).all()
    
    # Top 6 most-tracked products (by number of price entries)
    product_counts = db.session.query(
        Product.id, Product.name, db.func.count(PriceEntry.id).label('count')
    ).outerjoin(PriceEntry).group_by(Product.id).order_by(db.desc('count')).limit(6).all()
    
    return render_template('index.html',
                         total_contacts=total_contacts,
                         total_stores=total_stores,
                         total_products=total_products,
                         total_price_entries=total_price_entries,
                         recent_prices=recent_prices,
                         top_products=product_counts)

@app.route('/compare')
def compare():
    products = Product.query.order_by(Product.name).all()
    stores = Store.query.order_by(Store.name).all()
    selected_product_id = request.args.get('product_id', type=int)
    
    comparison_data = None
    price_history = None
    
    if selected_product_id:
        product = Product.query.get_or_404(selected_product_id)
        stores = Store.query.all()
        
        comparison_data = []
        for store in stores:
            latest_entry = PriceEntry.query.filter_by(
                product_id=selected_product_id, 
                store_id=store.id
            ).order_by(PriceEntry.date_checked.desc()).first()
            
            if latest_entry:
                comparison_data.append({
                    'store': store,
                    'price': latest_entry.price,
                    'date_checked': latest_entry.date_checked,
                    'notes': latest_entry.notes
                })
        
        # Sort by price (cheapest first)
        comparison_data.sort(key=lambda x: x['price'])
        
        # Add ranking info
        if comparison_data:
            min_price = comparison_data[0]['price']
            max_price = comparison_data[-1]['price']
            for item in comparison_data:
                item['savings'] = item['price'] - min_price
                item['is_cheapest'] = item['price'] == min_price
                item['is_most_expensive'] = item['price'] == max_price
        
        # Get price history for chart
        price_history = {}
        for store in stores:
            entries = PriceEntry.query.filter_by(
                product_id=selected_product_id,
                store_id=store.id
            ).order_by(PriceEntry.date_checked).all()
            
            if entries:
                price_history[store.name] = {
                    'dates': [e.date_checked.strftime('%Y-%m-%d') for e in entries],
                    'prices': [e.price for e in entries]
                }
    
    # Build full matrix for all products × all stores
    all_products = Product.query.order_by(Product.name).all()
    all_stores = Store.query.order_by(Product.name).all()
    
    matrix_data = []
    for product in all_products:
        row = {'product': product, 'stores': [], 'min_price': None, 'max_price': None}
        prices = []
        
        for store in all_stores:
            latest = PriceEntry.query.filter_by(
                product_id=product.id,
                store_id=store.id
            ).order_by(PriceEntry.date_checked.desc()).first()
            
            if latest:
                row['stores'].append({'store': store, 'price': latest.price, 'date': latest.date_checked})
                prices.append(latest.price)
            else:
                row['stores'].append({'store': store, 'price': None, 'date': None})
        
        if prices:
            row['min_price'] = min(prices)
            row['max_price'] = max(prices)
        
        matrix_data.append(row)
    
    return render_template('compare.html',
                         products=products,
                         selected_product_id=selected_product_id,
                         comparison_data=comparison_data,
                         price_history=price_history,
                         matrix_data=matrix_data,
                         all_stores=all_stores)

@app.route('/prices')
def prices():
    product_filter = request.args.get('product_id', type=int)
    store_filter = request.args.get('store_id', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = PriceEntry.query.order_by(PriceEntry.date_checked.desc())
    
    if product_filter:
        query = query.filter_by(product_id=product_filter)
    if store_filter:
        query = query.filter_by(store_id=store_filter)
    if date_from:
        query = query.filter(PriceEntry.date_checked >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(PriceEntry.date_checked <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    price_entries = query.all()
    products = Product.query.order_by(Product.name).all()
    stores = Store.query.order_by(Product.name).all()
    
    return render_template('prices.html',
                         price_entries=price_entries,
                         products=products,
                         stores=stores,
                         product_filter=product_filter,
                         store_filter=store_filter,
                         date_from=date_from,
                         date_to=date_to)

@app.route('/prices/add', methods=['GET', 'POST'])
def add_price():
    if request.method == 'POST':
        product_id = request.form.get('product_id', type=int)
        store_id = request.form.get('store_id', type=int)
        price = request.form.get('price', type=float)
        date_checked_str = request.form.get('date_checked')
        notes = request.form.get('notes', '')
        
        if not product_id or not store_id or not price:
            flash('Please fill in all required fields.', 'error')
            return redirect(url_for('add_price'))
        
        date_checked = datetime.strptime(date_checked_str, '%Y-%m-%d').date() if date_checked_str else date.today()
        
        new_entry = PriceEntry(
            product_id=product_id,
            store_id=store_id,
            price=price,
            date_checked=date_checked,
            notes=notes
        )
        
        db.session.add(new_entry)
        db.session.commit()
        
        flash('Price entry added successfully!', 'success')
        return redirect(url_for('prices'))
    
    products = Product.query.order_by(Product.name).all()
    stores = Store.query.order_by(Product.name).all()
    today = date.today().strftime('%Y-%m-%d')
    
    return render_template('price_form.html',
                         products=products,
                         stores=stores,
                         entry=None,
                         action='Add',
                         today=today)

@app.route('/prices/edit/<int:id>', methods=['GET', 'POST'])
def edit_price(id):
    entry = PriceEntry.query.get_or_404(id)
    
    if request.method == 'POST':
        entry.product_id = request.form.get('product_id', type=int)
        entry.store_id = request.form.get('store_id', type=int)
        entry.price = request.form.get('price', type=float)
        date_checked_str = request.form.get('date_checked')
        entry.date_checked = datetime.strptime(date_checked_str, '%Y-%m-%d').date() if date_checked_str else date.today()
        entry.notes = request.form.get('notes', '')
        
        db.session.commit()
        
        flash('Price entry updated successfully!', 'success')
        return redirect(url_for('prices'))
    
    products = Product.query.order_by(Product.name).all()
    stores = Store.query.order_by(Product.name).all()
    
    return render_template('price_form.html',
                         products=products,
                         stores=stores,
                         entry=entry,
                         action='Edit')

@app.route('/prices/delete/<int:id>', methods=['POST'])
def delete_price(id):
    entry = PriceEntry.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    
    flash('Price entry deleted successfully!', 'success')
    return redirect(url_for('prices'))

@app.route('/products')
def products():
    search = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    
    query = Product.query.order_by(Product.name)
    
    if search:
        query = query.filter(
            db.or_(
                Product.name.ilike(f'%{search}%'),
                Product.brand.ilike(f'%{search}%')
            )
        )
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    product_list = query.all()
    categories = ['Membrane', 'Coating', 'Primer', 'Sealant', 'Adhesive', 'Accessory']
    
    return render_template('products.html',
                         products=product_list,
                         categories=categories,
                         search=search,
                         category_filter=category_filter)

@app.route('/products/add', methods=['GET', 'POST'])
def add_product():
    if request.method == 'POST':
        name = request.form.get('name')
        unit = request.form.get('unit')
        category = request.form.get('category')
        brand = request.form.get('brand')
        description = request.form.get('description', '')
        
        if not name:
            flash('Product name is required.', 'error')
            return redirect(url_for('add_product'))
        
        new_product = Product(
            name=name,
            unit=unit,
            category=category,
            brand=brand,
            description=description
        )
        
        db.session.add(new_product)
        db.session.commit()
        
        flash('Product added successfully!', 'success')
        return redirect(url_for('products'))
    
    categories = ['Membrane', 'Coating', 'Primer', 'Sealant', 'Adhesive', 'Accessory']
    units = ['bag', 'liter', 'sqm', 'pcs', 'roll', 'set']
    
    return render_template('product_form.html',
                         categories=categories,
                         units=units,
                         product=None,
                         action='Add')

@app.route('/products/edit/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        product.name = request.form.get('name')
        product.unit = request.form.get('unit')
        product.category = request.form.get('category')
        product.brand = request.form.get('brand')
        product.description = request.form.get('description', '')
        
        db.session.commit()
        
        flash('Product updated successfully!', 'success')
        return redirect(url_for('products'))
    
    categories = ['Membrane', 'Coating', 'Primer', 'Sealant', 'Adhesive', 'Accessory']
    units = ['bag', 'liter', 'sqm', 'pcs', 'roll', 'set']
    
    return render_template('product_form.html',
                         categories=categories,
                         units=units,
                         product=product,
                         action='Edit')

@app.route('/products/delete/<int:id>', methods=['POST'])
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    
    flash('Product deleted successfully!', 'success')
    return redirect(url_for('products'))

@app.route('/stores')
def stores():
    store_list = Store.query.order_by(Store.name).all()
    return render_template('stores.html', stores=store_list)

@app.route('/stores/add', methods=['GET', 'POST'])
def add_store():
    if request.method == 'POST':
        name = request.form.get('name')
        location = request.form.get('location')
        contact_person = request.form.get('contact_person')
        phone = request.form.get('phone')
        notes = request.form.get('notes', '')
        
        if not name:
            flash('Store name is required.', 'error')
            return redirect(url_for('add_store'))
        
        new_store = Store(
            name=name,
            location=location,
            contact_person=contact_person,
            phone=phone,
            notes=notes
        )
        
        db.session.add(new_store)
        db.session.commit()
        
        flash('Store added successfully!', 'success')
        return redirect(url_for('stores'))
    
    return render_template('store_form.html', store=None, action='Add')

@app.route('/stores/edit/<int:id>', methods=['GET', 'POST'])
def edit_store(id):
    store = Store.query.get_or_404(id)
    
    if request.method == 'POST':
        store.name = request.form.get('name')
        store.location = request.form.get('location')
        store.contact_person = request.form.get('contact_person')
        store.phone = request.form.get('phone')
        store.notes = request.form.get('notes', '')
        
        db.session.commit()
        
        flash('Store updated successfully!', 'success')
        return redirect(url_for('stores'))
    
    return render_template('store_form.html', store=store, action='Edit')

@app.route('/stores/delete/<int:id>', methods=['POST'])
def delete_store(id):
    store = Store.query.get_or_404(id)
    db.session.delete(store)
    db.session.commit()
    
    flash('Store deleted successfully!', 'success')
    return redirect(url_for('stores'))

@app.route('/contacts')
def contacts():
    search = request.args.get('search', '')
    role_filter = request.args.get('role', '')
    
    query = Contact.query.order_by(Contact.name)
    
    if search:
        query = query.filter(
            db.or_(
                Contact.name.ilike(f'%{search}%'),
                Contact.company.ilike(f'%{search}%'),
                Contact.phone.ilike(f'%{search}%'),
                Contact.email.ilike(f'%{search}%')
            )
        )
    if role_filter:
        query = query.filter_by(role=role_filter)
    
    contact_list = query.all()
    roles = ['Supplier', 'Client', 'Contractor', 'Other']
    
    return render_template('contacts.html',
                         contacts=contact_list,
                         roles=roles,
                         search=search,
                         role_filter=role_filter)

@app.route('/contacts/add', methods=['GET', 'POST'])
def add_contact():
    if request.method == 'POST':
        name = request.form.get('name')
        company = request.form.get('company')
        role = request.form.get('role')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')
        notes = request.form.get('notes', '')
        
        if not name:
            flash('Contact name is required.', 'error')
            return redirect(url_for('add_contact'))
        
        new_contact = Contact(
            name=name,
            company=company,
            role=role,
            phone=phone,
            email=email,
            address=address,
            notes=notes
        )
        
        db.session.add(new_contact)
        db.session.commit()
        
        flash('Contact added successfully!', 'success')
        return redirect(url_for('contacts'))
    
    roles = ['Supplier', 'Client', 'Contractor', 'Other']
    
    return render_template('contact_form.html', roles=roles, contact=None, action='Add')

@app.route('/contacts/edit/<int:id>', methods=['GET', 'POST'])
def edit_contact(id):
    contact = Contact.query.get_or_404(id)
    
    if request.method == 'POST':
        contact.name = request.form.get('name')
        contact.company = request.form.get('company')
        contact.role = request.form.get('role')
        contact.phone = request.form.get('phone')
        contact.email = request.form.get('email')
        contact.address = request.form.get('address')
        contact.notes = request.form.get('notes', '')
        
        db.session.commit()
        
        flash('Contact updated successfully!', 'success')
        return redirect(url_for('contacts'))
    
    roles = ['Supplier', 'Client', 'Contractor', 'Other']
    
    return render_template('contact_form.html', roles=roles, contact=contact, action='Edit')

@app.route('/contacts/delete/<int:id>', methods=['POST'])
def delete_contact(id):
    contact = Contact.query.get_or_404(id)
    db.session.delete(contact)
    db.session.commit()
    
    flash('Contact deleted successfully!', 'success')
    return redirect(url_for('contacts'))

@app.route('/export/comparison')
def export_comparison():
    wb = Workbook()
    
    # Sheet 1: Price Comparison Matrix
    ws1 = wb.active
    ws1.title = "Price Comparison"
    
    # Headers
    headers = ['Product', 'Category', 'Unit']
    stores = Store.query.order_by(Store.name).all()
    for store in stores:
        headers.append(store.name)
    headers.extend(['Cheapest', 'Most Expensive', 'Price Range'])
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    green_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    red_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    light_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    products = Product.query.order_by(Product.name).all()
    for row_idx, product in enumerate(products, 2):
        prices = []
        
        ws1.cell(row=row_idx, column=1, value=product.name).border = thin_border
        ws1.cell(row=row_idx, column=2, value=product.category).border = thin_border
        ws1.cell(row=row_idx, column=3, value=product.unit).border = thin_border
        
        for col_idx, store in enumerate(stores, 4):
            latest = PriceEntry.query.filter_by(
                product_id=product.id,
                store_id=store.id
            ).order_by(PriceEntry.date_checked.desc()).first()
            
            if latest:
                cell = ws1.cell(row=row_idx, column=col_idx, value=latest.price)
                prices.append(latest.price)
            else:
                cell = ws1.cell(row=row_idx, column=col_idx, value='N/A')
            
            cell.border = thin_border
            cell.number_format = '₱#,##0.00'
        
        # Summary columns
        if prices:
            min_price = min(prices)
            max_price = max(prices)
            price_range = max_price - min_price
            
            # Find and highlight cheapest/most expensive
            for col_idx, price in enumerate(prices, 4):
                cell = ws1.cell(row=row_idx, column=col_idx)
                if price == min_price:
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif price == max_price:
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")
            
            ws1.cell(row=row_idx, column=len(headers)-2, value=min_price).border = thin_border
            ws1.cell(row=row_idx, column=len(headers)-2).number_format = '₱#,##0.00'
            ws1.cell(row=row_idx, column=len(headers)-1, value=max_price).border = thin_border
            ws1.cell(row=row_idx, column=len(headers)-1).number_format = '₱#,##0.00'
            ws1.cell(row=row_idx, column=len(headers), value=price_range).border = thin_border
            ws1.cell(row=row_idx, column=len(headers)).number_format = '₱#,##0.00'
        
        # Alternating row colors
        fill = light_fill if row_idx % 2 == 0 else white_fill
        for col in range(1, len(headers)+1):
            if not ws1.cell(row=row_idx, column=col).fill.start_color.rgb:
                ws1.cell(row=row_idx, column=col).fill = fill
    
    # Adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws1.column_dimensions[column].width = adjusted_width
    
    # Sheet 2: Price Log
    ws2 = wb.create_sheet(title="Price Log")
    log_headers = ['Date', 'Product', 'Store', 'Price', 'Notes']
    
    for col, header in enumerate(log_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    price_entries = PriceEntry.query.order_by(PriceEntry.date_checked.desc()).all()
    for row_idx, entry in enumerate(price_entries, 2):
        ws2.cell(row=row_idx, column=1, value=entry.date_checked.strftime('%Y-%m-%d')).border = thin_border
        ws2.cell(row=row_idx, column=2, value=entry.product.name).border = thin_border
        ws2.cell(row=row_idx, column=3, value=entry.store.name).border = thin_border
        cell = ws2.cell(row=row_idx, column=4, value=entry.price)
        cell.border = thin_border
        cell.number_format = '₱#,##0.00'
        ws2.cell(row=row_idx, column=5, value=entry.notes).border = thin_border
        
        fill = light_fill if row_idx % 2 == 0 else white_fill
        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).fill = fill
    
    # Sheet 3: Contacts
    ws3 = wb.create_sheet(title="Contacts")
    contact_headers = ['Name', 'Company', 'Role', 'Phone', 'Email', 'Address', 'Notes']
    
    for col, header in enumerate(contact_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    contacts = Contact.query.order_by(Contact.name).all()
    for row_idx, contact in enumerate(contacts, 2):
        ws3.cell(row=row_idx, column=1, value=contact.name).border = thin_border
        ws3.cell(row=row_idx, column=2, value=contact.company).border = thin_border
        ws3.cell(row=row_idx, column=3, value=contact.role).border = thin_border
        ws3.cell(row=row_idx, column=4, value=contact.phone).border = thin_border
        ws3.cell(row=row_idx, column=5, value=contact.email).border = thin_border
        ws3.cell(row=row_idx, column=6, value=contact.address).border = thin_border
        ws3.cell(row=row_idx, column=7, value=contact.notes).border = thin_border
        
        fill = light_fill if row_idx % 2 == 0 else white_fill
        for col in range(1, 8):
            ws3.cell(row=row_idx, column=col).fill = fill
    
    # Sheet 4: Stores
    ws4 = wb.create_sheet(title="Stores")
    store_headers = ['Name', 'Location', 'Contact Person', 'Phone', 'Notes']
    
    for col, header in enumerate(store_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    store_list = Store.query.order_by(Store.name).all()
    for row_idx, store in enumerate(store_list, 2):
        ws4.cell(row=row_idx, column=1, value=store.name).border = thin_border
        ws4.cell(row=row_idx, column=2, value=store.location).border = thin_border
        ws4.cell(row=row_idx, column=3, value=store.contact_person).border = thin_border
        ws4.cell(row=row_idx, column=4, value=store.phone).border = thin_border
        ws4.cell(row=row_idx, column=5, value=store.notes).border = thin_border
        
        fill = light_fill if row_idx % 2 == 0 else white_fill
        for col in range(1, 6):
            ws4.cell(row=row_idx, column=col).fill = fill
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"price_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/export/template')
def export_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    headers = ['product_name', 'unit', 'category', 'brand', 'store_name', 'store_location', 'price', 'date_checked', 'notes']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Example rows
    example1 = ['Waterproofing Membrane', 'sqm', 'Membrane', 'Brand A', 'Store X', 'Manila', 150.00, '2024-01-15', 'Good quality']
    example2 = ['Primer Coat', 'liter', 'Primer', 'Brand B', 'Store Y', 'Quezon City', 250.00, '2024-01-15', 'Fast drying']
    
    for row_idx, example in enumerate([example1, example2], 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='import_template.xlsx'
    )

@app.route('/import', methods=['GET', 'POST'])
def import_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded.', 'error')
            return redirect(url_for('import_data'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected.', 'error')
            return redirect(url_for('import_data'))
        
        if not file.filename.endswith('.xlsx'):
            flash('Please upload an Excel file (.xlsx).', 'error')
            return redirect(url_for('import_data'))
        
        try:
            df = pd.read_excel(file)
            
            required_cols = ['product_name', 'unit', 'store_name', 'price', 'date_checked']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                flash(f'Missing required columns: {", ".join(missing_cols)}', 'error')
                return redirect(url_for('import_data'))
            
            success_count = 0
            error_messages = []
            
            for idx, row in df.iterrows():
                try:
                    # Get or create product
                    product = Product.query.filter_by(name=row['product_name']).first()
                    if not product:
                        product = Product(
                            name=row['product_name'],
                            unit=row.get('unit', 'pcs'),
                            category=row.get('category', ''),
                            brand=row.get('brand', ''),
                            description=''
                        )
                        db.session.add(product)
                        db.session.flush()
                    
                    # Get or create store
                    store = Store.query.filter_by(name=row['store_name']).first()
                    if not store:
                        store = Store(
                            name=row['store_name'],
                            location=row.get('store_location', ''),
                            contact_person='',
                            phone='',
                            notes=''
                        )
                        db.session.add(store)
                        db.session.flush()
                    
                    # Create price entry
                    date_checked = row['date_checked']
                    if isinstance(date_checked, str):
                        date_checked = datetime.strptime(date_checked, '%Y-%m-%d').date()
                    elif hasattr(date_checked, 'date'):
                        date_checked = date_checked.date()
                    
                    price_entry = PriceEntry(
                        product_id=product.id,
                        store_id=store.id,
                        price=float(row['price']),
                        date_checked=date_checked,
                        notes=row.get('notes', '')
                    )
                    db.session.add(price_entry)
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"Row {idx + 2}: {str(e)}")
            
            db.session.commit()
            
            if error_messages:
                flash(f'Imported {success_count} entries with {len(error_messages)} errors.', 'warning')
                for err in error_messages[:5]:
                    flash(err, 'error')
            else:
                flash(f'Successfully imported {success_count} price entries!', 'success')
            
            return redirect(url_for('prices'))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return redirect(url_for('import_data'))
    
    return render_template('import.html')

@app.route('/api/history/<int:product_id>')
def api_history(product_id):
    stores = Store.query.all()
    history = {}
    
    for store in stores:
        entries = PriceEntry.query.filter_by(
            product_id=product_id,
            store_id=store.id
        ).order_by(PriceEntry.date_checked).all()
        
        if entries:
            history[store.name] = {
                'dates': [e.date_checked.strftime('%Y-%m-%d') for e in entries],
                'prices': [e.price for e in entries]
            }
    
    return jsonify(history)

@app.route('/api/matrix')
def api_matrix():
    products = Product.query.order_by(Product.name).all()
    stores = Store.query.order_by(Product.name).all()
    
    matrix = []
    for product in products:
        row = {
            'product_id': product.id,
            'product_name': product.name,
            'prices': {}
        }
        
        for store in stores:
            latest = PriceEntry.query.filter_by(
                product_id=product.id,
                store_id=store.id
            ).order_by(PriceEntry.date_checked.desc()).first()
            
            if latest:
                row['prices'][store.name] = {
                    'price': latest.price,
                    'date': latest.date_checked.strftime('%Y-%m-%d')
                }
        
        matrix.append(row)
    
    return jsonify(matrix)

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
